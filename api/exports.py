"""api/exports.py — Report export generators (v5.3-05).

Generates PDF, CSV, and XLSX exports from stored report JSON.
PDF requires weasyprint + markdown (optional dependencies).
CSV uses stdlib. XLSX uses openpyxl.
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def generate_csv(report: dict[str, Any]) -> str:
    """Extract key metrics from a report as a flat CSV string."""
    meta = report.get("meta", {})
    scoring = report.get("scoring", {})
    cashflow = report.get("cashflow", {})
    debt = report.get("debt", {})
    goals = report.get("goals", {})
    investments = report.get("investments", {})
    mortgage = report.get("mortgage", {})

    surplus = cashflow.get("surplus", {})
    net_income = cashflow.get("net_income", {})
    savings_rate = cashflow.get("savings_rate", {})
    debt_summary = debt.get("summary", {})
    goal_summary = goals.get("summary", {})
    pension = investments.get("pension_analysis", {})
    portfolio = investments.get("current_portfolio", {})

    rows = [
        ("Section", "Metric", "Value"),
        ("Profile", "Name", meta.get("profile_name")),
        ("Profile", "Generated At", meta.get("generated_at")),
        ("Profile", "Engine Version", meta.get("engine_version")),
        ("Scoring", "Overall Score", scoring.get("overall_score")),
        ("Scoring", "Grade", scoring.get("grade")),
        ("Income", "Gross Annual", cashflow.get("income", {}).get("total_gross_annual")),
        ("Income", "Net Annual", net_income.get("annual")),
        ("Income", "Net Monthly", net_income.get("monthly")),
        ("Expenses", "Total Monthly", cashflow.get("expenses", {}).get("total_monthly")),
        ("Cashflow", "Surplus Monthly", surplus.get("monthly")),
        ("Cashflow", "Surplus Annual", surplus.get("annual")),
        ("Cashflow", "Savings Rate %", savings_rate.get("basic_pct")),
        ("Cashflow", "Savings Rate (incl Pension) %", savings_rate.get("effective_pct_incl_pension")),
        ("Debt", "Total Balance", debt_summary.get("total_balance")),
        ("Debt", "Total Minimum Monthly", debt_summary.get("total_minimum_monthly")),
        ("Debt", "DTI (Gross) %", debt_summary.get("debt_to_income_gross_pct")),
        ("Debt", "High Interest Count", debt_summary.get("high_interest_debt_count")),
        ("Debt", "Strategy", debt.get("recommended_strategy")),
        ("Goals", "Total", goal_summary.get("total_goals")),
        ("Goals", "On Track", goal_summary.get("on_track")),
        ("Goals", "At Risk", goal_summary.get("at_risk")),
        ("Goals", "Unreachable", goal_summary.get("unreachable")),
        ("Investments", "Total Invested", portfolio.get("total_invested")),
        ("Investments", "ISA Balance", portfolio.get("isa_balance")),
        ("Investments", "Pension Balance", portfolio.get("pension_balance")),
        ("Investments", "Pension Projected (Real)", pension.get("projected_at_retirement_real")),
        ("Investments", "Income Replacement %", pension.get("income_replacement_ratio_pct")),
        ("Mortgage", "Readiness", _get_mortgage_readiness(mortgage)),
    ]

    # Add scoring category breakdown
    for cat_name, cat_data in scoring.get("categories", {}).items():
        if isinstance(cat_data, dict):
            rows.append(("Scoring", f"{cat_name} Score", cat_data.get("score")))

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def generate_xlsx(report: dict[str, Any]) -> bytes:
    """Generate a multi-sheet Excel workbook from a report."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    _build_summary_sheet(wb, report)
    _build_cashflow_sheet(wb, report)
    _build_debt_sheet(wb, report)
    _build_goals_sheet(wb, report)
    _build_investments_sheet(wb, report)

    # Style all sheet headers
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(wb, report: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Summary"

    meta = report.get("meta", {})
    scoring = report.get("scoring", {})
    cashflow = report.get("cashflow", {})
    surplus = cashflow.get("surplus", {})
    debt_summary = report.get("debt", {}).get("summary", {})
    goal_summary = report.get("goals", {}).get("summary", {})
    portfolio = report.get("investments", {}).get("current_portfolio", {})

    ws.append(["Metric", "Value"])
    ws.append(["Profile Name", meta.get("profile_name")])
    ws.append(["Generated At", meta.get("generated_at")])
    ws.append(["Overall Score", scoring.get("overall_score")])
    ws.append(["Grade", scoring.get("grade")])
    ws.append(["Net Income (Monthly)", cashflow.get("net_income", {}).get("monthly")])
    ws.append(["Surplus (Monthly)", surplus.get("monthly")])
    ws.append(["Savings Rate %", cashflow.get("savings_rate", {}).get("basic_pct")])
    ws.append(["Total Debt", debt_summary.get("total_balance")])
    ws.append(["Goals On Track", goal_summary.get("on_track")])
    ws.append(["Goals At Risk", goal_summary.get("at_risk")])
    ws.append(["Total Invested", portfolio.get("total_invested")])
    ws.append([])

    # Scoring categories
    ws.append(["Scoring Category", "Score", "Weight"])
    for cat_name, cat_data in scoring.get("categories", {}).items():
        if isinstance(cat_data, dict):
            ws.append([cat_name.replace("_", " ").title(), cat_data.get("score"), cat_data.get("weight")])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12


def _build_cashflow_sheet(wb, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Cashflow")
    cashflow = report.get("cashflow", {})
    income = cashflow.get("income", {})
    deductions = cashflow.get("deductions", {})
    expenses = cashflow.get("expenses", {})
    surplus = cashflow.get("surplus", {})

    ws.append(["Category", "Annual", "Monthly"])
    ws.append(["Gross Income", income.get("total_gross_annual"), income.get("total_gross_monthly")])
    ws.append(["Income Tax", deductions.get("income_tax_annual"), _safe_div(deductions.get("income_tax_annual"), 12)])
    ws.append(["National Insurance", deductions.get("national_insurance_annual"), _safe_div(deductions.get("national_insurance_annual"), 12)])
    ws.append(["Pension (Personal)", deductions.get("pension_personal_annual"), _safe_div(deductions.get("pension_personal_annual"), 12)])
    ws.append(["Net Income", cashflow.get("net_income", {}).get("annual"), cashflow.get("net_income", {}).get("monthly")])
    ws.append(["Expenses", expenses.get("total_annual"), expenses.get("total_monthly")])
    ws.append(["Debt Servicing", cashflow.get("debt_servicing", {}).get("total_annual"), cashflow.get("debt_servicing", {}).get("total_monthly")])
    ws.append(["Surplus", surplus.get("annual"), surplus.get("monthly")])

    # Expense category breakdown
    breakdown = expenses.get("category_breakdown_monthly", {})
    if breakdown:
        ws.append([])
        ws.append(["Expense Category", "Monthly"])
        for cat, amount in sorted(breakdown.items()):
            ws.append([cat.replace("_", " ").title(), amount])

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16


def _build_debt_sheet(wb, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Debt")
    debt = report.get("debt", {})
    debts = debt.get("debts", [])

    ws.append(["Name", "Type", "Balance", "Rate %", "Min Payment", "Payoff Months"])
    for d in debts:
        if isinstance(d, dict):
            ws.append([
                d.get("name"),
                d.get("type"),
                d.get("balance"),
                d.get("interest_rate"),
                d.get("minimum_payment_monthly"),
                d.get("months_to_payoff"),
            ])

    # Summary row
    summary = debt.get("summary", {})
    if summary:
        ws.append([])
        ws.append(["Total", "", summary.get("total_balance"), "", summary.get("total_minimum_monthly")])

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 16


def _build_goals_sheet(wb, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Goals")
    goals_list = report.get("goals", {}).get("goals", [])

    ws.append(["Name", "Target", "Deadline (Years)", "Priority", "Status", "Monthly Needed", "Feasibility"])
    for g in goals_list:
        if isinstance(g, dict):
            ws.append([
                g.get("name"),
                g.get("target_amount"),
                g.get("deadline_years"),
                g.get("priority"),
                g.get("status"),
                g.get("monthly_contribution_needed"),
                g.get("feasibility"),
            ])

    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 18


def _build_investments_sheet(wb, report: dict[str, Any]) -> None:
    ws = wb.create_sheet("Investments")
    inv = report.get("investments", {})
    portfolio = inv.get("current_portfolio", {})
    pension = inv.get("pension_analysis", {})
    allocation = inv.get("suggested_allocation", {})

    ws.append(["Portfolio Summary", "Value"])
    ws.append(["ISA Balance", portfolio.get("isa_balance")])
    ws.append(["LISA Balance", portfolio.get("lisa_balance")])
    ws.append(["Pension Balance", portfolio.get("pension_balance")])
    ws.append(["Other Investments", portfolio.get("other_investments")])
    ws.append(["Total Invested", portfolio.get("total_invested")])
    ws.append([])

    ws.append(["Pension Projection", "Value"])
    ws.append(["Monthly Contribution", pension.get("monthly_contribution_total")])
    ws.append(["Projected at Retirement (Nominal)", pension.get("projected_at_retirement_nominal")])
    ws.append(["Projected at Retirement (Real)", pension.get("projected_at_retirement_real")])
    ws.append(["Income Replacement %", pension.get("income_replacement_ratio_pct")])
    ws.append([])

    if isinstance(allocation, dict) and allocation:
        ws.append(["Suggested Allocation", "Weight %"])
        for asset, weight in allocation.items():
            ws.append([asset.replace("_", " ").title(), weight])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def generate_pdf(report: dict[str, Any]) -> bytes:
    """Generate a PDF from the report's narrative Markdown.

    Requires `markdown` and `weasyprint` packages.
    Raises ImportError if either is missing.
    """
    import markdown
    import weasyprint

    from engine.narrative import generate_narrative

    narrative_md = generate_narrative(report)
    html_body = markdown.markdown(narrative_md, extensions=["tables", "fenced_code"])

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 40px; line-height: 1.6; color: #333; font-size: 11pt; }}
h1 {{ color: #1a5276; border-bottom: 2px solid #1a5276; padding-bottom: 8px; font-size: 20pt; }}
h2 {{ color: #2c3e50; margin-top: 24px; font-size: 15pt; }}
h3 {{ color: #34495e; font-size: 12pt; }}
table {{ border-collapse: collapse; width: 100%; margin: 12px 0; }}
th, td {{ border: 1px solid #bdc3c7; padding: 8px 12px; text-align: left; }}
th {{ background-color: #ecf0f1; font-weight: bold; }}
tr:nth-child(even) {{ background-color: #f8f9fa; }}
.score {{ font-size: 28pt; font-weight: bold; color: #1a5276; }}
hr {{ border: none; border-top: 1px solid #bdc3c7; margin: 20px 0; }}
</style></head><body>{html_body}</body></html>"""

    pdf_bytes = weasyprint.HTML(string=html).write_pdf()
    return pdf_bytes


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_mortgage_readiness(mortgage: dict[str, Any]) -> str | None:
    """Extract mortgage readiness from various possible structures."""
    if isinstance(mortgage, dict):
        if "readiness" in mortgage:
            r = mortgage["readiness"]
            return r.get("status") if isinstance(r, dict) else r
        return mortgage.get("applicable")
    return None


def _safe_div(value: float | int | None, divisor: float) -> float | None:
    if value is None:
        return None
    return round(value / divisor, 2)

"""Tests for report export generators and endpoints (v5.3-05)."""

from __future__ import annotations

import csv
import io
import json
import os

os.environ["DATABASE_URL"] = "sqlite://"

from unittest.mock import patch

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from api.database.models import Base, Run
from api.database.session import get_db

_test_engine = create_engine(
    "sqlite://",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
_TestSession = sessionmaker(bind=_test_engine, autoflush=False, expire_on_commit=False)

Base.metadata.create_all(bind=_test_engine)


def _override_get_db():
    session = _TestSession()
    try:
        yield session
    finally:
        session.close()


from api.main import app

app.dependency_overrides[get_db] = _override_get_db

from fastapi.testclient import TestClient

client = TestClient(app)

API_KEY = "dev-key-change-me"
HEADERS = {"X-API-Key": API_KEY}


def _sample_report() -> dict:
    """Minimal report structure for export tests."""
    return {
        "meta": {
            "profile_name": "Export Test User",
            "generated_at": "2026-04-09T10:00:00",
            "engine_version": "5.3.5",
        },
        "scoring": {
            "overall_score": 65.0,
            "grade": "B-",
            "categories": {
                "savings_rate": {"score": 70.0, "weight": 0.2, "detail": "Good"},
                "debt_health": {"score": 55.0, "weight": 0.2, "detail": "Fair"},
            },
        },
        "cashflow": {
            "income": {"total_gross_annual": 50000, "total_gross_monthly": 4166.67},
            "deductions": {"income_tax_annual": 7486, "national_insurance_annual": 4964, "pension_personal_annual": 2500},
            "net_income": {"annual": 35050, "monthly": 2920.83},
            "expenses": {
                "total_monthly": 2000,
                "total_annual": 24000,
                "category_breakdown_monthly": {"housing": 1000, "transport": 300, "living": 500, "other": 200},
            },
            "debt_servicing": {"total_monthly": 100, "total_annual": 1200},
            "surplus": {"monthly": 820.83, "annual": 9850},
            "savings_rate": {"basic_pct": 15.0, "effective_pct_incl_pension": 22.0},
        },
        "debt": {
            "debts": [
                {"name": "Credit Card", "type": "credit_card", "balance": 2000, "interest_rate": 19.9, "minimum_payment_monthly": 50, "months_to_payoff": 60},
            ],
            "summary": {
                "total_balance": 2000,
                "total_minimum_monthly": 50,
                "debt_to_income_gross_pct": 4.0,
                "high_interest_debt_count": 1,
            },
            "recommended_strategy": "avalanche",
            "avalanche_order": [],
            "extra_payment_scenarios": [],
        },
        "goals": {
            "goals": [
                {"name": "Emergency Fund", "target_amount": 10000, "deadline_years": 2, "priority": 1, "status": "on_track", "monthly_contribution_needed": 200, "feasibility": "feasible"},
            ],
            "prerequisites": {},
            "summary": {"total_goals": 1, "on_track": 1, "at_risk": 0, "unreachable": 0, "blocked": 0},
        },
        "investments": {
            "current_portfolio": {"isa_balance": 5000, "lisa_balance": 0, "pension_balance": 10000, "other_investments": 0, "total_invested": 15000},
            "pension_analysis": {
                "monthly_contribution_total": 208.33,
                "projected_at_retirement_nominal": 500000,
                "projected_at_retirement_real": 250000,
                "income_replacement_ratio_pct": 45.0,
            },
            "suggested_allocation": {"uk_equity": 40, "global_equity": 30, "government_bonds": 20, "corporate_bonds": 10},
            "risk_profile": "moderate",
        },
        "mortgage": {"readiness": "not_ready"},
        "life_events": {},
        "insurance": {},
        "stress_scenarios": {},
        "estate": {},
        "sensitivity_analysis": {},
        "advisor_insights": {},
        "review_schedule": {},
        "validation": {"flags": [], "error_count": 0, "warning_count": 0, "info_count": 0},
    }


def _create_run_in_db() -> int:
    """Insert a run with sample report JSON and return its ID."""
    session = _TestSession()
    try:
        run = Run(
            timestamp="2026-04-09T10:00:00",
            profile_name="Export Test User",
            overall_score=65.0,
            grade="B-",
            full_report_json=json.dumps(_sample_report()),
        )
        session.add(run)
        session.commit()
        session.refresh(run)
        return run.id
    finally:
        session.close()


# ---------------------------------------------------------------------------
# CSV generator unit tests
# ---------------------------------------------------------------------------

class TestCsvGenerator:
    def test_generates_valid_csv(self):
        from api.exports import generate_csv

        result = generate_csv(_sample_report())
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        assert rows[0] == ["Section", "Metric", "Value"]
        assert len(rows) > 10

    def test_includes_key_metrics(self):
        from api.exports import generate_csv

        result = generate_csv(_sample_report())
        assert "Export Test User" in result
        assert "65.0" in result
        assert "B-" in result
        assert "820.83" in result

    def test_includes_scoring_categories(self):
        from api.exports import generate_csv

        result = generate_csv(_sample_report())
        assert "savings_rate" in result
        assert "debt_health" in result

    def test_handles_empty_report(self):
        from api.exports import generate_csv

        result = generate_csv({})
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        assert rows[0] == ["Section", "Metric", "Value"]


# ---------------------------------------------------------------------------
# XLSX generator unit tests
# ---------------------------------------------------------------------------

class TestXlsxGenerator:
    def test_generates_valid_xlsx(self):
        from api.exports import generate_xlsx

        result = generate_xlsx(_sample_report())
        assert isinstance(result, bytes)
        assert len(result) > 0
        # XLSX files start with PK (ZIP header)
        assert result[:2] == b"PK"

    def test_has_expected_sheets(self):
        import openpyxl

        from api.exports import generate_xlsx

        result = generate_xlsx(_sample_report())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Cashflow" in sheet_names
        assert "Debt" in sheet_names
        assert "Goals" in sheet_names
        assert "Investments" in sheet_names

    def test_summary_sheet_has_score(self):
        import openpyxl

        from api.exports import generate_xlsx

        result = generate_xlsx(_sample_report())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert 65.0 in values
        assert "B-" in values

    def test_debt_sheet_has_debts(self):
        import openpyxl

        from api.exports import generate_xlsx

        result = generate_xlsx(_sample_report())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Debt"]
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert "Credit Card" in values

    def test_handles_empty_report(self):
        from api.exports import generate_xlsx

        result = generate_xlsx({})
        assert isinstance(result, bytes)
        assert len(result) > 0


# ---------------------------------------------------------------------------
# API endpoint tests
# ---------------------------------------------------------------------------

class TestExportEndpoints:
    def test_csv_export_endpoint(self):
        run_id = _create_run_in_db()
        resp = client.get(f"/api/v1/export/{run_id}/csv", headers=HEADERS)
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "text/csv; charset=utf-8"
        assert "Export Test User" in resp.text

    def test_xlsx_export_endpoint(self):
        run_id = _create_run_in_db()
        resp = client.get(f"/api/v1/export/{run_id}/xlsx", headers=HEADERS)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert resp.content[:2] == b"PK"

    def test_pdf_export_missing_deps(self):
        run_id = _create_run_in_db()
        with patch("api.exports.generate_pdf", side_effect=ImportError("no weasyprint")):
            resp = client.get(f"/api/v1/export/{run_id}/pdf", headers=HEADERS)
            assert resp.status_code == 501

    def test_export_nonexistent_run_returns_404(self):
        resp = client.get("/api/v1/export/99999/csv", headers=HEADERS)
        assert resp.status_code == 404

    def test_export_allowed_in_dev_mode(self):
        resp = client.get("/api/v1/export/1/csv")
        # Dev mode allows unauthenticated access; 404 because run doesn't exist
        assert resp.status_code in (200, 404)

    def test_csv_content_disposition(self):
        run_id = _create_run_in_db()
        resp = client.get(f"/api/v1/export/{run_id}/csv", headers=HEADERS)
        assert "attachment" in resp.headers.get("content-disposition", "")
        assert ".csv" in resp.headers.get("content-disposition", "")

    def test_xlsx_content_disposition(self):
        run_id = _create_run_in_db()
        resp = client.get(f"/api/v1/export/{run_id}/xlsx", headers=HEADERS)
        assert "attachment" in resp.headers.get("content-disposition", "")
        assert ".xlsx" in resp.headers.get("content-disposition", "")
